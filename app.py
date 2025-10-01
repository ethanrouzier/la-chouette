from flask import Flask, render_template, request, jsonify, send_file
import os
import json
import pandas as pd
from werkzeug.utils import secure_filename
import PyPDF2
import docx
from PIL import Image
import pytesseract
import uuid
import numpy as np
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from sklearn.feature_extraction.text import TfidfVectorizer
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import base64
from io import BytesIO

# Imports pour le style Excel
try:
    from openpyxl.styles import PatternFill, Font
    OPENPYXL_STYLES_AVAILABLE = True
except ImportError:
    OPENPYXL_STYLES_AVAILABLE = False

# Variables pour les imports optionnels
SENTENCE_TRANSFORMERS_AVAILABLE = False
MISTRAL_AVAILABLE = False

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['JSON_FOLDER'] = 'documents_json'
app.config['CATALOG_FILE'] = 'catalog.json'
app.config['API_KEY_FILE'] = 'mistral_api_key.txt'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Créer les dossiers nécessaires
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['JSON_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'txt', 'pdf', 'docx', 'png', 'jpg', 'jpeg', 'xlsx', 'xls', 'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_pdf(file_path):
    """Extrait le texte d'un fichier PDF"""
    try:
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text()
        return text
    except Exception as e:
        return f"Erreur lors de l'extraction du PDF: {str(e)}"

def extract_text_from_docx(file_path):
    """Extrait le texte d'un fichier Word"""
    try:
        doc = docx.Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        return f"Erreur lors de l'extraction du Word: {str(e)}"

def extract_text_from_image(file_path):
    """Extrait le texte d'une image avec OCR"""
    try:
        image = Image.open(file_path)
        text = pytesseract.image_to_string(image, lang='fra')
        return text
    except Exception as e:
        return f"Erreur lors de l'extraction OCR: {str(e)}"

def process_document(file_path, filename):
    """Traite un document et retourne le titre et le contenu"""
    file_ext = filename.rsplit('.', 1)[1].lower()
    
    if file_ext == 'txt':
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
    elif file_ext == 'pdf':
        content = extract_text_from_pdf(file_path)
    elif file_ext == 'docx':
        content = extract_text_from_docx(file_path)
    elif file_ext in ['png', 'jpg', 'jpeg']:
        content = extract_text_from_image(file_path)
    else:
        content = "Format non supporté"
    
    # Utiliser le nom du fichier comme titre par défaut
    title = filename.rsplit('.', 1)[0]
    
    return title, content

def perform_clustering(texts, doc_ids, n_docs):
    """Effectue le clustering des documents"""
    if n_docs < 2:
        return [{'cluster_id': 0, 'documents': doc_ids, 'name': 'Tous les documents'}]
    
    # Essayer d'importer sentence-transformers dynamiquement
    try:
        from sentence_transformers import SentenceTransformer
        model = SentenceTransformer('sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2')
        embeddings = model.encode(texts, convert_to_numpy=True, normalize_embeddings=True)
    except Exception as e:
        print(f"Erreur embeddings: {e}")
        return perform_simple_clustering(texts, doc_ids, n_docs)
    
    # Déterminer le nombre optimal de clusters
    max_k = min(10, n_docs // 2)
    min_k = 2
    
    if max_k < min_k:
        return [{'cluster_id': 0, 'documents': doc_ids, 'name': 'Tous les documents'}]
    
    best_k = min_k
    best_score = -1
    
    for k in range(min_k, max_k + 1):
        try:
            kmeans = KMeans(n_clusters=k, random_state=42, n_init=10)
            labels = kmeans.fit_predict(embeddings)
            
            if len(set(labels)) > 1:
                score = silhouette_score(embeddings, labels)
                if score > best_score:
                    best_score = score
                    best_k = k
        except Exception:
            continue
    
    # Clustering final
    kmeans = KMeans(n_clusters=best_k, random_state=42, n_init=10)
    labels = kmeans.fit_predict(embeddings)
    
    # Organiser les résultats
    clusters = {}
    for i, (doc_id, label) in enumerate(zip(doc_ids, labels)):
        if label not in clusters:
            clusters[label] = []
        clusters[label].append(doc_id)
    
    # Créer la structure de retour
    result = []
    for cluster_id, documents in clusters.items():
        result.append({
            'cluster_id': int(cluster_id),
            'documents': documents,
            'name': f'Cluster {cluster_id + 1}'
        })
    
    return result

def perform_simple_clustering(texts, doc_ids, n_docs):
    """Clustering simple basé sur TF-IDF"""
    try:
        # Utiliser TF-IDF pour vectoriser les textes
        vectorizer = TfidfVectorizer(max_features=1000, stop_words=None)
        tfidf_matrix = vectorizer.fit_transform(texts)
        
        # Déterminer le nombre optimal de clusters
        max_k = min(5, n_docs // 2)
        min_k = 2
        
        if max_k < min_k:
            return [{'cluster_id': 0, 'documents': doc_ids, 'name': 'Tous les documents'}]
        
        best_k = min_k
        best_score = -1
        
        for k in range(min_k, max_k + 1):
            try:
                kmeans = KMeans(n_clusters=k, random_state=42, n_init=10)
                labels = kmeans.fit_predict(tfidf_matrix.toarray())
                
                if len(set(labels)) > 1:
                    score = silhouette_score(tfidf_matrix.toarray(), labels)
                    if score > best_score:
                        best_score = score
                        best_k = k
            except Exception:
                continue
        
        # Clustering final
        kmeans = KMeans(n_clusters=best_k, random_state=42, n_init=10)
        labels = kmeans.fit_predict(tfidf_matrix.toarray())
        
        # Organiser les résultats
        clusters = {}
        for i, (doc_id, label) in enumerate(zip(doc_ids, labels)):
            if label not in clusters:
                clusters[label] = []
            clusters[label].append(doc_id)
        
        # Créer la structure de retour
        result = []
        for cluster_id, documents in clusters.items():
            result.append({
                'cluster_id': int(cluster_id),
                'documents': documents,
                'name': f'Cluster {cluster_id + 1}'
            })
        
        return result
        
    except Exception as e:
        print(f"Erreur clustering simple: {e}")
        return [{'cluster_id': 0, 'documents': doc_ids, 'name': 'Tous les documents'}]

def name_clusters_with_mistral(clusters, api_key, instructions):
    """Nomme les clusters avec Mistral"""
    try:
        from mistralai import Mistral
        client = Mistral(api_key=api_key)
    except ImportError:
        raise Exception("Mistral AI n'est pas disponible. Veuillez installer mistralai.")
    
    try:
        for cluster in clusters:
            # Récupérer les documents du cluster
            doc_ids = cluster['documents']
            documents = []
            
            for doc_id in doc_ids:
                json_path = os.path.join(app.config['JSON_FOLDER'], f"{doc_id}.json")
                if os.path.exists(json_path):
                    with open(json_path, 'r', encoding='utf-8') as file:
                        doc_data = json.load(file)
                        documents.append({
                            'title': doc_data.get('title', ''),
                            'content': doc_data.get('content', '')[:1000]  # Limiter la taille
                        })
            
            if not documents:
                continue
            
            # Préparer le prompt
            samples_text = ""
            for i, doc in enumerate(documents[:5]):  # Max 5 documents
                samples_text += f"Document {i+1}:\n"
                samples_text += f"Titre: {doc['title']}\n"
                samples_text += f"Contenu: {doc['content'][:500]}...\n\n"
            
            system_prompt = (
                "Tu es un expert en catégorisation de documents. "
                "Analyse les documents suivants et donne UNIQUEMENT le nom de la catégorie "
                "en 1-4 mots clairs et précis. "
                "Réponds uniquement avec le nom de la catégorie, sans explication."
            )
            
            if instructions:
                system_prompt += f"\n\nConsignes spécifiques: {instructions}"
            
            user_prompt = f"Documents du cluster:\n\n{samples_text}"
            
            # Appel à Mistral
            response = client.chat.complete(
                model="mistral-large-latest",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.2,
                max_tokens=50
            )
            
            cluster_name = response.choices[0].message.content.strip()
            # Nettoyer le nom
            cluster_name = cluster_name.replace('"', '').replace("'", '').strip()
            if not cluster_name:
                cluster_name = f"Cluster {cluster['cluster_id'] + 1}"
            
            cluster['name'] = cluster_name
            
            # Mettre à jour les documents avec la catégorie
            for doc_id in doc_ids:
                json_path = os.path.join(app.config['JSON_FOLDER'], f"{doc_id}.json")
                if os.path.exists(json_path):
                    with open(json_path, 'r', encoding='utf-8') as file:
                        doc_data = json.load(file)
                    
                    doc_data['category'] = cluster_name
                    
                    with open(json_path, 'w', encoding='utf-8') as file:
                        json.dump(doc_data, file, ensure_ascii=False, indent=2)
        
        return clusters
        
    except Exception as e:
        print(f"Erreur Mistral: {e}")
        raise e

def create_cluster_visualization(documents):
    """Crée une visualisation des clusters"""
    try:
        # Préparer les données
        categories = {}
        for doc in documents:
            category = doc.get('category', 'Non catégorisé')
            if category not in categories:
                categories[category] = []
            categories[category].append(doc)
        
        # Créer le graphique
        fig, ax = plt.subplots(figsize=(10, 6))
        
        category_names = list(categories.keys())
        category_counts = [len(docs) for docs in categories.values()]
        
        # Couleurs
        colors = plt.cm.Set3(np.linspace(0, 1, len(category_names)))
        
        bars = ax.bar(category_names, category_counts, color=colors)
        
        # Personnaliser le graphique
        ax.set_title('Répartition des documents par catégorie', fontsize=14, fontweight='bold')
        ax.set_xlabel('Catégories', fontsize=12)
        ax.set_ylabel('Nombre de documents', fontsize=12)
        
        # Rotation des labels
        plt.xticks(rotation=45, ha='right')
        
        # Ajouter les valeurs sur les barres
        for bar, count in zip(bars, category_counts):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                   f'{count}', ha='center', va='bottom')
        
        plt.tight_layout()
        
        # Convertir en base64
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        buffer.seek(0)
        image_base64 = base64.b64encode(buffer.getvalue()).decode()
        plt.close()
        
        return image_base64
        
    except Exception as e:
        print(f"Erreur visualisation: {e}")
        return None

def generate_fields_with_ai(category, documents, num_fields, existing_fields, api_key=None, instructions=None):
    """Génère des champs pour une catégorie avec Mistral AI"""
    try:
        # Si pas de clé API, utiliser la génération basique
        if not api_key:
            return generate_fields_basic(category, documents, num_fields, existing_fields)
        
        # Appel à Mistral AI
        from mistralai import Mistral
        client = Mistral(api_key=api_key)
        
        # Préparer les exemples (1-2 documents)
        examples = []
        for doc in documents[:2]:
            content = f"{doc.get('title', '')} {doc.get('content', '')}"
            # Tronquer le contenu pour éviter les tokens excessifs
            content = content[:2000] + "..." if len(content) > 2000 else content
            examples.append(content)
        
        # Construire le prompt
        system_prompt = (
            "Tu es un expert en extraction de données et structuration de documents. "
            "Ton objectif est de proposer des champs pertinents pour catégoriser et structurer des documents. "
            "Réponds UNIQUEMENT en JSON, sans Markdown ni explications. "
            "Format attendu: [{\"name\": \"nom_du_champ\", \"type\": \"text|number|date|boolean\", \"description\": \"Description du champ\"}]"
        )
        
        user_prompt = f"""
Catégorie: {category}
Nombre de champs souhaités: {num_fields}

Exemples de documents:
{chr(10).join([f"Document {i+1}: {ex}" for i, ex in enumerate(examples)])}

Champs existants à réutiliser si pertinents: {', '.join(existing_fields[:20])}

Instructions spécifiques: {instructions or 'Aucune'}

Propose {num_fields} champs pertinents pour cette catégorie de documents.
Inclus toujours un champ 'date_doc' de type 'date'.
Réponds uniquement avec le JSON des champs.
"""
        
        # Appel à Mistral
        response = client.chat.complete(
            model="mistral-large-latest",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.2,
            max_tokens=1000
        )
        
        # Parser la réponse JSON
        response_text = response.choices[0].message.content.strip()
        
        # Nettoyer la réponse (enlever markdown si présent)
        if "```json" in response_text:
            response_text = response_text.split("```json")[1].split("```")[0]
        elif "```" in response_text:
            response_text = response_text.split("```")[1].split("```")[0]
        
        # Parser le JSON
        import json
        try:
            fields_data = json.loads(response_text)
            if isinstance(fields_data, list):
                return fields_data[:num_fields]
            else:
                raise ValueError("Format de réponse invalide")
        except json.JSONDecodeError:
            # Fallback vers génération basique
            return generate_fields_basic(category, documents, num_fields, existing_fields)
        
    except Exception as e:
        print(f"Erreur génération champs Mistral: {e}")
        # Fallback vers génération basique
        return generate_fields_basic(category, documents, num_fields, existing_fields)

def generate_fields_basic(category, documents, num_fields, existing_fields):
    """Génération basique de champs (fallback)"""
    # Générer des champs basés sur le contexte
    if 'médical' in category.lower() or 'medical' in category.lower():
        suggested_fields = [
            {'name': 'diagnostic', 'type': 'text', 'description': 'Diagnostic médical'},
            {'name': 'traitement', 'type': 'text', 'description': 'Traitement prescrit'},
            {'name': 'date_consultation', 'type': 'date', 'description': 'Date de consultation'},
            {'name': 'medecin', 'type': 'text', 'description': 'Nom du médecin'},
            {'name': 'patient', 'type': 'text', 'description': 'Informations patient'},
            {'name': 'medicaments', 'type': 'text', 'description': 'Médicaments prescrits'},
            {'name': 'symptomes', 'type': 'text', 'description': 'Symptômes observés'}
        ]
    elif 'juridique' in category.lower() or 'legal' in category.lower():
        suggested_fields = [
            {'name': 'type_contrat', 'type': 'text', 'description': 'Type de contrat'},
            {'name': 'parties', 'type': 'text', 'description': 'Parties impliquées'},
            {'name': 'date_signature', 'type': 'date', 'description': 'Date de signature'},
            {'name': 'montant', 'type': 'number', 'description': 'Montant financier'},
            {'name': 'duree', 'type': 'text', 'description': 'Durée du contrat'},
            {'name': 'clauses', 'type': 'text', 'description': 'Clauses importantes'},
            {'name': 'juridiction', 'type': 'text', 'description': 'Juridiction compétente'}
        ]
    elif 'technique' in category.lower() or 'technical' in category.lower():
        suggested_fields = [
            {'name': 'version', 'type': 'text', 'description': 'Version du logiciel'},
            {'name': 'fonctionnalite', 'type': 'text', 'description': 'Fonctionnalité décrite'},
            {'name': 'langage', 'type': 'text', 'description': 'Langage de programmation'},
            {'name': 'environnement', 'type': 'text', 'description': 'Environnement d\'exécution'},
            {'name': 'dependances', 'type': 'text', 'description': 'Dépendances requises'},
            {'name': 'exemple_code', 'type': 'text', 'description': 'Exemple de code'},
            {'name': 'documentation', 'type': 'text', 'description': 'Lien vers documentation'}
        ]
    else:
        # Génération générique basée sur le contenu
        suggested_fields = [
            {'name': 'titre', 'type': 'text', 'description': 'Titre du document'},
            {'name': 'date_doc', 'type': 'date', 'description': 'Date du document'},
            {'name': 'auteur', 'type': 'text', 'description': 'Auteur du document'},
            {'name': 'type', 'type': 'text', 'description': 'Type de document'},
            {'name': 'contenu_principal', 'type': 'text', 'description': 'Contenu principal'},
            {'name': 'mots_cles', 'type': 'text', 'description': 'Mots-clés'},
            {'name': 'statut', 'type': 'text', 'description': 'Statut du document'}
        ]
    
    # Limiter au nombre demandé
    suggested_fields = suggested_fields[:num_fields]
    
    # Ajouter des champs existants si disponibles
    for field in existing_fields:
        if field not in [f['name'] for f in suggested_fields]:
            suggested_fields.append({
                'name': field,
                'type': 'text',
                'description': f'Champ existant: {field}'
            })
    
    return suggested_fields

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    """Gère l'upload des fichiers"""
    if 'files' not in request.files:
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    
    files = request.files.getlist('files')
    uploaded_documents = []
    
    for file in files:
        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # Traiter le document
            title, content = process_document(file_path, filename)
            
            # Créer un ID unique pour le document
            doc_id = str(uuid.uuid4())
            
            # Sauvegarder en JSON
            json_data = {
                'id': doc_id,
                'title': title,
                'content': content,
                'filename': filename,
                'type': 'document'
            }
            
            json_path = os.path.join(app.config['JSON_FOLDER'], f"{doc_id}.json")
            with open(json_path, 'w', encoding='utf-8') as json_file:
                json.dump(json_data, json_file, ensure_ascii=False, indent=2)
            
            uploaded_documents.append({
                'id': doc_id,
                'title': title,
                'filename': filename,
                'type': 'document'
            })
    
    return jsonify({'documents': uploaded_documents})

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    """Gère l'upload et le traitement des fichiers Excel/CSV"""
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    
    file = request.files['file']
    text_column = request.form.get('text_column')
    title_column = request.form.get('title_column', '')
    start_row = int(request.form.get('start_row', 1))
    end_row = int(request.form.get('end_row', 0))
    
    if file and file.filename and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        try:
            # Lire le fichier Excel ou CSV
            if filename.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
            
            # Ajuster les indices pour commencer à 0
            start_idx = start_row - 1
            end_idx = end_row if end_row > 0 else len(df)
            
            # Filtrer les lignes
            df_filtered = df.iloc[start_idx:end_idx]
            
            created_documents = []
            
            for index, row in df_filtered.iterrows():
                # Récupérer le contenu texte
                content = str(row[text_column]) if text_column in df.columns else ""
                
                # Récupérer le titre
                if title_column and title_column in df.columns:
                    title = str(row[title_column])
                else:
                    title = f"Row {index + 1}"
                
                # Créer un ID unique
                doc_id = str(uuid.uuid4())
                
                # Sauvegarder en JSON
                json_data = {
                    'id': doc_id,
                    'title': title,
                    'content': content,
                    'filename': filename,
                    'row_index': index + 1,
                    'type': 'excel_row'
                }
                
                json_path = os.path.join(app.config['JSON_FOLDER'], f"{doc_id}.json")
                with open(json_path, 'w', encoding='utf-8') as json_file:
                    json.dump(json_data, json_file, ensure_ascii=False, indent=2)
                
                created_documents.append({
                    'id': doc_id,
                    'title': title,
                    'content': content[:100] + "..." if len(content) > 100 else content,
                    'row_index': index + 1
                })
            
            return jsonify({
                'success': True,
                'documents': created_documents,
                'total_rows': len(df_filtered)
            })
            
        except Exception as e:
            return jsonify({'error': f'Erreur lors du traitement du fichier: {str(e)}'}), 400
    
    return jsonify({'error': 'Fichier non valide'}), 400

@app.route('/get_excel_columns', methods=['POST'])
def get_excel_columns():
    """Récupère les colonnes d'un fichier Excel/CSV"""
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier sélectionné'}), 400
    
    file = request.files['file']
    if file and file.filename and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_{filename}")
        file.save(file_path)
        
        try:
            if filename.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
            
            # Nettoyer le fichier temporaire
            os.remove(file_path)
            
            return jsonify({
                'columns': list(df.columns),
                'total_rows': len(df)
            })
            
        except Exception as e:
            return jsonify({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}), 400
    
    return jsonify({'error': 'Fichier non valide'}), 400

@app.route('/get_document/<doc_id>')
def get_document(doc_id):
    """Récupère le contenu d'un document"""
    json_path = os.path.join(app.config['JSON_FOLDER'], f"{doc_id}.json")
    
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        return jsonify(data)
    
    return jsonify({'error': 'Document non trouvé'}), 404

@app.route('/categorization')
def categorization():
    """Page de catégorisation"""
    return render_template('categorization.html')

@app.route('/get_all_documents')
def get_all_documents():
    """Récupère tous les documents pour la catégorisation"""
    documents = []
    json_folder = app.config['JSON_FOLDER']
    
    if os.path.exists(json_folder):
        for filename in os.listdir(json_folder):
            if filename.endswith('.json'):
                json_path = os.path.join(json_folder, filename)
                try:
                    with open(json_path, 'r', encoding='utf-8') as file:
                        data = json.load(file)
                        documents.append(data)
                except Exception as e:
                    print(f"Erreur lecture {filename}: {e}")
    
    return jsonify({'documents': documents})

@app.route('/cluster_documents', methods=['POST'])
def cluster_documents():
    """Effectue le clustering automatique des documents"""
    try:
        data = request.get_json()
        documents = data.get('documents', [])
        # Utiliser la clé API stockée ou celle fournie
        api_key = data.get('api_key') or load_api_key()
        instructions = data.get('instructions', '')
        
        if not documents:
            return jsonify({'error': 'Aucun document à traiter'}), 400
        
        # Préparer les données
        texts = []
        doc_ids = []
        for doc in documents:
            text = f"{doc.get('title', '')} {doc.get('content', '')}"
            texts.append(text)
            doc_ids.append(doc['id'])
        
        # Clustering
        clustering_result = perform_clustering(texts, doc_ids, len(documents))
        
        # Nommage avec Mistral si clé API fournie
        if api_key:
            try:
                clustering_result = name_clusters_with_mistral(
                    clustering_result, api_key, instructions
                )
            except Exception as e:
                return jsonify({
                    'error': f'Erreur lors du nommage Mistral: {str(e)}',
                    'clusters': clustering_result
                }), 400
        
        return jsonify({'clusters': clustering_result})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors du clustering: {str(e)}'}), 500

@app.route('/update_document_category', methods=['POST'])
def update_document_category():
    """Met à jour la catégorie d'un document"""
    try:
        data = request.get_json()
        doc_id = data.get('doc_id')
        category = data.get('category')
        
        if not doc_id or not category:
            return jsonify({'error': 'ID document et catégorie requis'}), 400
        
        json_path = os.path.join(app.config['JSON_FOLDER'], f"{doc_id}.json")
        
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as file:
                doc_data = json.load(file)
            
            doc_data['category'] = category
            
            with open(json_path, 'w', encoding='utf-8') as file:
                json.dump(doc_data, file, ensure_ascii=False, indent=2)
            
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Document non trouvé'}), 404
            
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la mise à jour: {str(e)}'}), 500

@app.route('/organize_documents_with_ai', methods=['POST'])
def organize_documents_with_ai():
    """Organise les documents non catégorisés dans les catégories existantes avec IA"""
    try:
        data = request.get_json()
        documents = data.get('documents', [])
        categories = data.get('categories', [])
        
        # Utiliser la clé API stockée côté serveur
        api_key = load_api_key()
        
        if not documents:
            return jsonify({'error': 'Aucun document à organiser'}), 400
        
        if not categories:
            return jsonify({'error': 'Aucune catégorie disponible'}), 400
        
        if not api_key:
            return jsonify({'error': 'Clé API Mistral requise. Veuillez la configurer dans les paramètres.'}), 400
        
        organized_documents = []
        
        for doc in documents:
            try:
                # Construire le prompt pour classifier le document
                prompt = f"""Tu es un expert en classification de documents. 

Voici un document à classer :
Titre : {doc.get('title', 'Sans titre')}
Contenu : {doc.get('content', '')[:1000]}...

Voici les catégories disponibles :
{', '.join(categories)}

À quelle catégorie ce document appartient-il ? Réponds UNIQUEMENT avec le nom exact de la catégorie, rien d'autre."""
                
                # Appeler l'API Mistral
                category = call_mistral_api(prompt, api_key)
                
                # Nettoyer et valider la réponse
                if category:
                    category = category.strip()
                    # Enlever les guillemets, points, etc.
                    category = category.strip('"').strip("'").strip('.').strip()
                    
                    # Vérifier que la catégorie retournée existe dans la liste (recherche insensible à la casse)
                    matched_category = None
                    for cat in categories:
                        if cat.lower() == category.lower() or category.lower() in cat.lower() or cat.lower() in category.lower():
                            matched_category = cat
                            break
                    
                    if matched_category:
                        organized_documents.append({
                            'id': doc['id'],
                            'category': matched_category
                        })
                        update_document_category_in_db(doc['id'], matched_category)
                    else:
                        # Si aucune correspondance exacte, essayer une correspondance partielle
                        print(f"Catégorie '{category}' non trouvée dans {categories}. Tentative de correspondance partielle...")
                        # Pour l'instant, laisser le document non catégorisé plutôt que de le mettre dans la première catégorie
                        organized_documents.append({
                            'id': doc['id'],
                            'category': None
                        })
                else:
                    print(f"Erreur API pour le document {doc['id']}. Document laissé non catégorisé.")
                    # Laisser le document non catégorisé plutôt que de le mettre dans la première catégorie
                    organized_documents.append({
                        'id': doc['id'],
                        'category': None
                    })
                    
            except Exception as e:
                print(f"Erreur lors de la classification du document {doc['id']}: {str(e)}")
                # En cas d'erreur, laisser le document non catégorisé
                organized_documents.append({
                    'id': doc['id'],
                    'category': None
                })
        
        return jsonify({
            'success': True,
            'organized_documents': organized_documents
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de l\'organisation par IA: {str(e)}'}), 500

def call_mistral_api(prompt, api_key):
    """Appelle l'API Mistral avec un prompt simple"""
    try:
        from mistralai import Mistral
        client = Mistral(api_key=api_key)
        
        response = client.chat.complete(
            model="mistral-large-latest",
            messages=[
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=100
        )
        
        return response.choices[0].message.content.strip()
        
    except Exception as e:
        print(f"Erreur lors de l'appel à Mistral: {str(e)}")
        return None

def update_document_category_in_db(doc_id, category):
    """Met à jour la catégorie d'un document dans la base de données"""
    try:
        json_path = os.path.join(app.config['JSON_FOLDER'], f"{doc_id}.json")
        
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as file:
                doc_data = json.load(file)
            
            doc_data['category'] = category
            
            with open(json_path, 'w', encoding='utf-8') as file:
                json.dump(doc_data, file, ensure_ascii=False, indent=2)
            
            return True
        return False
    except Exception as e:
        print(f"Erreur lors de la mise à jour de la catégorie du document {doc_id}: {str(e)}")
        return False

@app.route('/get_cluster_visualization')
def get_cluster_visualization():
    """Génère et retourne la visualisation des clusters"""
    try:
        # Récupérer tous les documents
        documents = []
        json_folder = app.config['JSON_FOLDER']
        
        if os.path.exists(json_folder):
            for filename in os.listdir(json_folder):
                if filename.endswith('.json'):
                    json_path = os.path.join(json_folder, filename)
                    try:
                        with open(json_path, 'r', encoding='utf-8') as file:
                            data = json.load(file)
                            if 'category' in data:
                                documents.append(data)
                    except Exception:
                        continue
        
        if len(documents) < 2:
            return jsonify({'error': 'Pas assez de documents catégorisés'}), 400
        
        # Créer la visualisation
        img_base64 = create_cluster_visualization(documents)
        
        return jsonify({'visualization': img_base64})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la création de la visualisation: {str(e)}'}), 500

@app.route('/field-selection')
def field_selection():
    """Page de sélection des champs"""
    return render_template('field_selection.html')

@app.route('/get_categories')
def get_categories():
    """Récupère toutes les catégories avec leurs documents"""
    try:
        categories = {}
        json_folder = app.config['JSON_FOLDER']
        
        if os.path.exists(json_folder):
            for filename in os.listdir(json_folder):
                if filename.endswith('.json'):
                    json_path = os.path.join(json_folder, filename)
                    try:
                        with open(json_path, 'r', encoding='utf-8') as file:
                            data = json.load(file)
                            if 'category' in data:
                                category = data['category']
                                if category not in categories:
                                    categories[category] = []
                                categories[category].append(data)
                    except Exception:
                        continue
        
        return jsonify({'categories': categories})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors du chargement des catégories: {str(e)}'}), 500

@app.route('/generate_fields', methods=['POST'])
def generate_fields():
    """Génère des champs pour une catégorie avec IA"""
    try:
        data = request.get_json()
        category = data.get('category')
        documents = data.get('documents', [])
        num_fields = data.get('num_fields', 7)
        existing_fields = data.get('existing_fields', [])
        # Utiliser la clé API stockée ou celle fournie
        api_key = data.get('api_key') or load_api_key()
        instructions = data.get('instructions')
        
        if not category or not documents:
            return jsonify({'error': 'Catégorie et documents requis'}), 400
        
        # Générer les champs avec IA
        generated_fields = generate_fields_with_ai(
            category, documents, num_fields, existing_fields, api_key, instructions
        )
        
        return jsonify({'fields': generated_fields})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la génération des champs: {str(e)}'}), 500

@app.route('/save_catalog', methods=['POST'])
def save_catalog():
    """Sauvegarde le catalog avec les champs et valeurs autorisées"""
    try:
        data = request.get_json()
        catalog = data.get('catalog', {})
        
        # Sauvegarder le catalog
        catalog_path = app.config['CATALOG_FILE']
        with open(catalog_path, 'w', encoding='utf-8') as file:
            json.dump(catalog, file, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la sauvegarde: {str(e)}'}), 500

@app.route('/get_catalog')
def get_catalog():
    """Récupère le catalog existant"""
    try:
        catalog_path = app.config['CATALOG_FILE']
        
        if os.path.exists(catalog_path):
            with open(catalog_path, 'r', encoding='utf-8') as file:
                catalog = json.load(file)
        else:
            catalog = {}
        
        return jsonify({'catalog': catalog})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors du chargement du catalog: {str(e)}'}), 500

@app.route('/generate_all_fields', methods=['POST'])
def generate_all_fields():
    """Génère des champs pour toutes les catégories avec IA"""
    try:
        data = request.get_json()
        categories = data.get('categories', {})
        num_fields = data.get('num_fields', 7)
        # Utiliser la clé API stockée ou celle fournie
        api_key = data.get('api_key') or load_api_key()
        instructions = data.get('instructions')
        
        if not categories:
            return jsonify({'error': 'Aucune catégorie fournie'}), 400
        
        results = {}
        existing_fields = []
        
        # Traiter chaque catégorie
        for category_name, documents in categories.items():
            try:
                # Générer les champs pour cette catégorie
                generated_fields = generate_fields_with_ai(
                    category_name, documents, num_fields, existing_fields, api_key, instructions
                )
                
                results[category_name] = {
                    'success': True,
                    'fields': generated_fields,
                    'count': len(generated_fields)
                }
                
                # Ajouter les nouveaux champs à la liste des champs existants
                for field in generated_fields:
                    if field.get('name') not in existing_fields:
                        existing_fields.append(field.get('name'))
                
            except Exception as e:
                results[category_name] = {
                    'success': False,
                    'error': str(e),
                    'fields': [],
                    'count': 0
                }
        
        return jsonify({
            'results': results,
            'total_categories': len(categories),
            'successful_categories': len([r for r in results.values() if r['success']])
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la génération en masse: {str(e)}'}), 500

@app.route('/generate_field_descriptions', methods=['POST'])
def generate_field_descriptions():
    """Génère des descriptions par IA pour les champs sans description"""
    try:
        data = request.get_json()
        fields = data.get('fields', [])
        catalog = data.get('catalog', {})
        
        # Utiliser la clé API stockée
        api_key = load_api_key()
        
        if not api_key:
            return jsonify({'error': 'Clé API Mistral requise. Veuillez la configurer dans les paramètres.'}), 400
        
        if not fields:
            return jsonify({'error': 'Aucun champ à traiter'}), 400
        
        descriptions = []
        
        for field in fields:
            try:
                category = field['category']
                field_name = field['fieldName']
                field_type = field['type']
                
                # Construire le prompt pour générer la description
                prompt = f"""Tu es un expert en structuration de données. 

Génère une description claire et concise pour le champ suivant :

Nom du champ : {field_name}
Type : {field_type}
Catégorie : {category}

Réponds UNIQUEMENT avec la description du champ en maximum 10 mots. La description doit être claire, précise et aider à comprendre à quoi sert ce champ. Pas d'explication ni de formatage."""
                
                # Appeler l'API Mistral
                description = call_mistral_api(prompt, api_key)
                
                if description:
                    description = description.strip()
                    # Nettoyer la réponse
                    description = description.strip('"').strip("'").strip()
                    
                    descriptions.append({
                        'category': category,
                        'fieldName': field_name,
                        'description': description
                    })
                else:
                    # Fallback : description basique
                    descriptions.append({
                        'category': category,
                        'fieldName': field_name,
                        'description': f"Champ {field_type} pour {category}"
                    })
                    
            except Exception as e:
                print(f"Erreur lors de la génération de description pour {field['fieldName']}: {str(e)}")
                # Fallback : description basique
                descriptions.append({
                    'category': field['category'],
                    'fieldName': field['fieldName'],
                    'description': f"Champ {field['type']} pour {field['category']}"
                })
        
        return jsonify({
            'success': True,
            'descriptions': descriptions,
            'count': len(descriptions)
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la génération des descriptions: {str(e)}'}), 500

@app.route('/extraction')
def extraction():
    """Page d'extraction des champs"""
    return render_template('extraction.html')

@app.route('/extract_fields', methods=['POST'])
def extract_fields():
    """Extrait les champs des documents avec Mistral AI"""
    try:
        data = request.get_json()
        documents = data.get('documents', [])
        catalog = data.get('catalog', {})
        field_descriptions = data.get('field_descriptions', {})
        # Utiliser la clé API stockée ou celle fournie
        api_key = data.get('api_key') or load_api_key()
        instructions = data.get('instructions', '')
        
        if not api_key:
            return jsonify({'error': 'Clé API Mistral requise. Veuillez la configurer dans les paramètres.'}), 400
        
        if not documents:
            return jsonify({'error': 'Aucun document à traiter'}), 400
        
        if not catalog:
            return jsonify({'error': 'Aucun catalog trouvé'}), 400
        
        # Traiter chaque document
        results = {}
        processed_documents = 0
        total_fields_extracted = 0
        
        for i, document in enumerate(documents):
            try:
                category = document.get('category', 'Non catégorisé')
                if category not in catalog:
                    continue
                
                # Extraire les champs pour cette catégorie
                extracted_fields = extract_document_fields(
                    document, 
                    catalog[category], 
                    field_descriptions.get(category, {}),
                    api_key, 
                    instructions
                )
                
                # Mettre à jour le document
                document['extracted_fields'] = extracted_fields
                
                # Sauvegarder le document mis à jour
                save_document_with_extracted_fields(document)
                
                results[document['id']] = {
                    'success': True,
                    'document_title': document.get('title', ''),
                    'extracted_fields': extracted_fields,
                    'field_count': len(extracted_fields)
                }
                
                processed_documents += 1
                total_fields_extracted += len(extracted_fields)
                
            except Exception as e:
                results[document['id']] = {
                    'success': False,
                    'error': str(e),
                    'document_title': document.get('title', ''),
                    'extracted_fields': {},
                    'field_count': 0
                }
        
        return jsonify({
            'results': results,
            'total_documents': len(documents),
            'processed_documents': processed_documents,
            'total_fields_extracted': total_fields_extracted
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de l\'extraction: {str(e)}'}), 500

def extract_document_fields(document, category_fields, field_descriptions, api_key, instructions):
    """Extrait les champs d'un document avec Mistral AI"""
    try:
        from mistralai import Mistral
        client = Mistral(api_key=api_key)
        
        extracted_fields = {}
        
        # Traiter chaque champ de la catégorie
        for field_name, field_config in category_fields.items():
            try:
                # Construire le prompt pour ce champ
                field_description = field_descriptions.get(field_name, field_config.get('description', ''))
                allowed_values = field_config.get('allowed_values', [])
                
                system_prompt = (
                    "Tu es un expert en extraction de données. "
                    "Ton objectif est d'extraire une valeur spécifique d'un document. "
                    "Réponds UNIQUEMENT avec la valeur extraite, sans explications ni formatage. "
                    "Si la valeur n'est pas trouvée, réponds 'N/A'."
                )
                
                user_prompt = f"""
Document: {document.get('title', '')} - {document.get('content', '')[:2000]}

Champ à extraire: {field_name}
Description: {field_description}
Type: {field_config.get('type', 'text')}
Valeurs autorisées: {', '.join(allowed_values) if allowed_values else 'Aucune'}

Instructions: {instructions}

Extrais la valeur pour le champ "{field_name}".
"""
                
                # Appel à Mistral
                response = client.chat.complete(
                    model="mistral-large-latest",
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt}
                    ],
                    temperature=0.1,
                    max_tokens=100
                )
                
                extracted_value = response.choices[0].message.content.strip()
                
                # Nettoyer la valeur
                if extracted_value.lower() in ['n/a', 'non trouvé', 'non disponible', '']:
                    extracted_value = None
                else:
                    # Vérifier si la valeur est dans les valeurs autorisées
                    if allowed_values and extracted_value not in allowed_values:
                        # Essayer de trouver une correspondance proche
                        for allowed_value in allowed_values:
                            if allowed_value.lower() in extracted_value.lower() or extracted_value.lower() in allowed_value.lower():
                                extracted_value = allowed_value
                                break
                
                extracted_fields[field_name] = extracted_value
                
            except Exception as e:
                print(f"Erreur extraction champ {field_name}: {e}")
                extracted_fields[field_name] = None
        
        return extracted_fields
        
    except Exception as e:
        print(f"Erreur extraction document: {e}")
        return {}

def save_document_with_extracted_fields(document):
    """Sauvegarde le document avec les champs extraits"""
    try:
        json_path = os.path.join(app.config['JSON_FOLDER'], f"{document['id']}.json")
        
        with open(json_path, 'w', encoding='utf-8') as file:
            json.dump(document, file, ensure_ascii=False, indent=2)
            
    except Exception as e:
        print(f"Erreur sauvegarde document {document['id']}: {e}")

def save_api_key(api_key):
    """Sauvegarde la clé API Mistral"""
    try:
        with open(app.config['API_KEY_FILE'], 'w', encoding='utf-8') as file:
            file.write(api_key)
        return True
    except Exception as e:
        print(f"Erreur sauvegarde clé API: {e}")
        return False

def load_api_key():
    """Charge la clé API Mistral"""
    try:
        if os.path.exists(app.config['API_KEY_FILE']):
            with open(app.config['API_KEY_FILE'], 'r', encoding='utf-8') as file:
                return file.read().strip()
        return None
    except Exception as e:
        print(f"Erreur chargement clé API: {e}")
        return None

@app.route('/api_key', methods=['GET', 'POST'])
def api_key():
    """Gestion de la clé API Mistral"""
    if request.method == 'GET':
        # Retourner la clé API (masquée)
        api_key = load_api_key()
        if api_key:
            masked_key = api_key[:8] + '*' * (len(api_key) - 12) + api_key[-4:] if len(api_key) > 12 else '***'
            return jsonify({'api_key': masked_key, 'has_key': True})
        else:
            return jsonify({'api_key': None, 'has_key': False})
    
    elif request.method == 'POST':
        # Sauvegarder la clé API
        data = request.get_json()
        api_key = data.get('api_key', '').strip()
        
        if not api_key:
            return jsonify({'error': 'Clé API requise'}), 400
        
        if save_api_key(api_key):
            return jsonify({'success': True, 'message': 'Clé API sauvegardée'})
        else:
            return jsonify({'error': 'Erreur lors de la sauvegarde'}), 500

@app.route('/validation')
def validation():
    """Page de validation des champs extraits"""
    return render_template('validation.html')

@app.route('/justify_field', methods=['POST'])
def justify_field():
    """Génère une justification IA pour un champ extrait"""
    try:
        data = request.get_json()
        document_content = data.get('document_content', '')
        field_name = data.get('field_name', '')
        field_value = data.get('field_value', '')
        
        # Utiliser la clé API stockée ou celle fournie
        api_key = data.get('api_key') or load_api_key()
        
        if not api_key:
            return jsonify({'error': 'Clé API Mistral requise. Veuillez la configurer dans les paramètres.'}), 400
        
        if not document_content or not field_name or not field_value:
            return jsonify({'error': 'Contenu du document, nom du champ et valeur requis'}), 400
        
        # Générer la justification avec Mistral AI
        justification = generate_field_justification(
            document_content, field_name, field_value, api_key
        )
        
        return jsonify(justification)
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la génération de la justification: {str(e)}'}), 500

def generate_field_justification(document_content, field_name, field_value, api_key):
    """Génère une justification IA pour un champ extrait"""
    try:
        from mistralai import Mistral
        client = Mistral(api_key=api_key)
        
        system_prompt = (
            "Tu es un expert en analyse de documents. "
            "Ton objectif est de trouver le passage exact dans un document qui justifie une valeur extraite. "
            "Réponds UNIQUEMENT avec le passage exact du document, sans explications ni formatage."
        )
        
        user_prompt = f"""
Document: {document_content[:3000]}

Champ extrait: {field_name}
Valeur extraite: {field_value}

Trouve le passage exact dans le document qui contient cette valeur. 
Réponds UNIQUEMENT avec le passage exact, sans explications.
"""
        
        response = client.chat.complete(
            model="mistral-large-latest",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1,
            max_tokens=500
        )
        
        response_text = response.choices[0].message.content.strip()
        
        # Nettoyer la réponse (enlever les guillemets et formatage)
        response_text = response_text.strip('"').strip("'").strip()
        
        # Si la réponse est vide ou trop courte, utiliser la valeur extraite
        if len(response_text) < 3:
            response_text = field_value
        
        # Sauvegarder la justification dans le document
        document_id = request.json.get('document_id')
        if document_id:
            document_path = os.path.join(app.config['JSON_FOLDER'], f'{document_id}.json')
            if os.path.exists(document_path):
                with open(document_path, 'r', encoding='utf-8') as file:
                    document_data = json.load(file)
                
                # Ajouter la justification au niveau 0
                if 'justifications' not in document_data:
                    document_data['justifications'] = {}
                
                document_data['justifications'][field_name] = {
                    "passage": response_text
                }
                
                with open(document_path, 'w', encoding='utf-8') as file:
                    json.dump(document_data, file, ensure_ascii=False, indent=2)
        
        return {
            "passage": response_text
        }
        
    except Exception as e:
        print(f"Erreur génération justification: {e}")
        return {
            "passage": field_value
        }

@app.route('/update_document', methods=['POST'])
def update_document():
    """Met à jour un document avec les modifications"""
    try:
        data = request.get_json()
        document = data.get('document')
        
        if not document or 'id' not in document:
            return jsonify({'error': 'Document invalide'}), 400
        
        # Sauvegarder le document mis à jour
        save_document_with_extracted_fields(document)
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la mise à jour: {str(e)}'}), 500

@app.route('/upload_catalog', methods=['POST'])
def upload_catalog():
    """Charge un catalog depuis l'ordinateur"""
    try:
        if 'catalog_file' not in request.files:
            return jsonify({'error': 'Aucun fichier fourni'}), 400
        
        file = request.files['catalog_file']
        if file.filename == '':
            return jsonify({'error': 'Aucun fichier sélectionné'}), 400
        
        if file and file.filename.endswith('.json'):
            # Lire le contenu du fichier
            content = file.read().decode('utf-8')
            catalog_data = json.loads(content)
            
            # Sauvegarder le catalog
            catalog_path = app.config['CATALOG_FILE']
            with open(catalog_path, 'w', encoding='utf-8') as f:
                json.dump(catalog_data, f, ensure_ascii=False, indent=2)
            
            return jsonify({'success': True, 'catalog': catalog_data})
        else:
            return jsonify({'error': 'Format de fichier non supporté. Utilisez un fichier .json'}), 400
        
    except json.JSONDecodeError:
        return jsonify({'error': 'Fichier JSON invalide'}), 400
    except Exception as e:
        return jsonify({'error': f'Erreur lors du chargement: {str(e)}'}), 500

@app.route('/download_catalog', methods=['GET'])
def download_catalog():
    """Télécharge le catalog en JSON"""
    try:
        catalog_path = app.config['CATALOG_FILE']
        
        if not os.path.exists(catalog_path):
            return jsonify({'error': 'Aucun catalog trouvé'}), 404
        
        with open(catalog_path, 'r', encoding='utf-8') as file:
            catalog_data = json.load(file)
        
        return jsonify(catalog_data)
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors du téléchargement: {str(e)}'}), 500

@app.route('/export_data', methods=['GET'])
def export_data():
    """Exporte les données extraites en Excel/CSV"""
    try:
        import pandas as pd
        from io import BytesIO
        import os
        
        format_type = request.args.get('format', 'excel').lower()
        
        # Charger tous les documents
        json_folder = app.config['JSON_FOLDER']
        documents = []
        
        if os.path.exists(json_folder):
            for filename in os.listdir(json_folder):
                if filename.endswith('.json'):
                    filepath = os.path.join(json_folder, filename)
                    with open(filepath, 'r', encoding='utf-8') as file:
                        doc_data = json.load(file)
                        documents.append(doc_data)
        
        if not documents:
            return jsonify({'error': 'Aucun document trouvé'}), 404
        
        # Créer le DataFrame
        rows = []
        extracted_field_names = set()
        justification_field_names = set()
        
        for doc in documents:
            row = {
                'Document ID': doc.get('id', ''),
                'Titre': doc.get('title', ''),
                'Catégorie': doc.get('category', ''),
                'Contenu': doc.get('content', '')[:100] + '...' if len(doc.get('content', '')) > 100 else doc.get('content', '')
            }
            
            # Ajouter les champs extraits
            extracted_fields = doc.get('extracted_fields', {})
            for field_name, field_value in extracted_fields.items():
                row[field_name] = field_value
                extracted_field_names.add(field_name)
            
            # Ajouter les justifications
            justifications = doc.get('justifications', {})
            for field_name, justification in justifications.items():
                justification_col_name = f"{field_name}_justification"
                row[justification_col_name] = justification.get('passage', '') if isinstance(justification, dict) else str(justification)
                justification_field_names.add(justification_col_name)
            
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        # Créer le fichier selon le format
        if format_type == 'csv':
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8')
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name='extracted_data.csv',
                mimetype='text/csv'
            )
        
        else:  # Excel avec couleurs
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Données Extraites')
                
                # Appliquer les couleurs aux titres des colonnes si disponible
                if OPENPYXL_STYLES_AVAILABLE:
                    worksheet = writer.sheets['Données Extraites']
                    
                    # Couleurs définies
                    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Vert clair
                    orange_fill = PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid')  # Orange clair
                    white_font = Font(color='FFFFFF')  # Police blanche pour contraste
                    
                    # Appliquer les couleurs aux en-têtes
                    for col_num, column_title in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        
                        if column_title in extracted_field_names:
                            # Champs extraits en vert
                            cell.fill = green_fill
                            cell.font = white_font
                        elif column_title in justification_field_names:
                            # Justifications en orange
                            cell.fill = orange_fill
                            cell.font = white_font
                        # Les autres colonnes (Document ID, Titre, Catégorie, Contenu) restent par défaut
            
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name='extracted_data.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de l\'export: {str(e)}'}), 500

@app.route('/reset_all', methods=['POST'])
def reset_all():
    """Réinitialise complètement l'application"""
    try:
        import shutil
        
        # Supprimer tous les documents JSON
        json_folder = app.config['JSON_FOLDER']
        if os.path.exists(json_folder):
            shutil.rmtree(json_folder)
            os.makedirs(json_folder, exist_ok=True)
        
        # Supprimer le catalog
        catalog_path = app.config['CATALOG_FILE']
        if os.path.exists(catalog_path):
            os.remove(catalog_path)
        
        # Supprimer les fichiers uploadés
        upload_folder = app.config['UPLOAD_FOLDER']
        if os.path.exists(upload_folder):
            shutil.rmtree(upload_folder)
            os.makedirs(upload_folder, exist_ok=True)
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la réinitialisation: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5001)
